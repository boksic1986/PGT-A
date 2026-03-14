#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# Author: jiucheng
# Updated: 2026/03/09
import argparse
import logging
import re
import shutil
from collections import defaultdict
from pathlib import Path

import yaml

from pipeline_logging import setup_logger

LOGGER = logging.getLogger("init_wes_project")


def detect_samples(fq_dir: Path) -> dict:
    fq_dir = Path(fq_dir)
    fq_files = (
        list(fq_dir.rglob("*.fastq.gz"))
        + list(fq_dir.rglob("*.fq.gz"))
        + list(fq_dir.rglob("*.fastq"))
        + list(fq_dir.rglob("*.fq"))
    )

    temp_dict = defaultdict(list)
    for fq in fq_files:
        name = fq.name
        base = re.sub(r"(?i)\.(fastq|fq)(\.gz)?$", "", name)
        match = re.search(r"(?i)(?:^|[._-])R?([12])(?:[._-]\d{3})?$", base)
        if not match:
            continue

        direction = "R1" if match.group(1) == "1" else "R2"
        sample_id = base[: match.start()].rstrip("._-")
        if not sample_id:
            continue
        temp_dict[sample_id].append((direction, fq))

    sample_dict = {}
    for sample_id, items in temp_dict.items():
        r1_list = sorted([fq for d, fq in items if d == "R1"])
        r2_list = sorted([fq for d, fq in items if d == "R2"])
        count = min(len(r1_list), len(r2_list))
        for i in range(count):
            key = f"{sample_id}_{i + 1}" if count > 1 else sample_id
            sample_dict[key] = {"R1": r1_list[i], "R2": r2_list[i]}
    return sample_dict


def load_template_config() -> dict:
    return {
        "core": {
            "project_path": "results",
            "reference_genome": "/data/Database/index/hg19/hg19.fa",
            "chromosome_list": [
                "chr1", "chr2", "chr3", "chr4", "chr5", "chr6", "chr7", "chr8", "chr9", "chr10",
                "chr11", "chr12", "chr13", "chr14", "chr15", "chr16", "chr17", "chr18", "chr19", "chr20",
                "chr21", "chr22", "chrX", "chrY",
            ],
            "wisecondorx": {
                "binsize": 100000,
                "reference_output": "wisecondorx/reference/ref_input.npz",
                "reference_output_by_sex": {
                    "XX": "reference/XX/result/ref_xx_best.npz",
                    "XY": "reference/XY/result/ref_xy_best.npz",
                },
                "reference_model_root": "reference",
                "use_chr_prefix": True,
                "tuning": {
                    "enable": True,
                    "bin_sizes": [100000, 200000, 300000, 500000, 750000, 1000000],
                    "pca": {"min_components": 2, "max_components": 20, "min_explained_variance": 0.0},
                    "qc": {
                        "min_reference_samples": 8,
                        "max_outlier_fraction": 0.25,
                        "min_reads_per_sample": 3000000,
                        "min_corr_to_median": 0.9,
                        "max_reconstruction_error_z": 3.5,
                        "max_noise_mad_z": 3.5,
                    },
                },
                "reference_prefilter": {
                    "binsize": 100000,
                    "max_iterations": 3,
                },
                "cnv": {
                    "enable": True,
                    "output_dir": "wisecondorx/cnv",
                    "qc_dir": "wisecondorx/cnv/qc",
                    "predict_dir": "wisecondorx/cnv/predict",
                    "convert_binsize": 100000,
                    "zscore": 5,
                    "alpha": 0.001,
                    "maskrepeats": 5,
                    "minrefbins": 150,
                    "qc": {"min_total_counts": 1000000, "min_nonzero_fraction": 0.4, "max_mad_log1p": 1.2},
                },
            },
        },
        "pipeline": {"targets": ["mapping", "reference", "cnv"]},
        "biosoft": {
            "fastp": "/biosoftware/bin/fastp",
            "bwa": "/biosoftware/bin/bwa",
            "samtools": "/biosoftware/bin/samtools",
            "WisecondorX": "/biosoftware/miniconda/envs/wise_env/bin/WisecondorX",
            "python": "/biosoftware/miniconda/envs/snakemake_env/bin/python",
        },
        "samples": {},
    }


def parse_id_tokens(text: str):
    normalized = (
        str(text)
        .strip()
        .replace("\u2014", "-")
        .replace("\u2013", "-")
        .replace("\uff0c", ",")
        .replace(" ", "")
    )
    tokens = []
    for item in normalized.split(","):
        if not item:
            continue
        if "-" in item:
            left, right = item.split("-", 1)
            if left.isdigit() and right.isdigit():
                a, b = int(left), int(right)
                if b < a:
                    a, b = b, a
                tokens.extend([str(x) for x in range(a, b + 1)])
                continue
            if re.fullmatch(r"[A-Za-z]", left) and re.fullmatch(r"[A-Za-z]", right):
                a, b = ord(left.upper()), ord(right.upper())
                if b < a:
                    a, b = b, a
                tokens.extend([chr(x) for x in range(a, b + 1)])
                continue
        tokens.append(item.upper() if re.fullmatch(r"[A-Za-z]+", item) else item)
    return tokens


def parse_reference_rules_from_excel(sample_info_xlsx: Path):
    import openpyxl

    wb = openpyxl.load_workbook(sample_info_xlsx, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise ValueError(f"Empty excel file: {sample_info_xlsx}")

    header = [str(x).strip() if x is not None else "" for x in rows[0]]
    col_map = {str(h).strip().lower(): idx for idx, h in enumerate(header)}

    required_cols = ["batch", "sample_ids", "sample_type", "sex"]
    missing = [col for col in required_cols if col not in col_map]
    if missing:
        raise ValueError(
            "Excel is missing required columns: "
            + ",".join(missing)
            + f". Header={header}"
        )

    idx_batch = col_map["batch"]
    idx_list = col_map["sample_ids"]
    idx_type = col_map["sample_type"]
    idx_sex = col_map["sex"]

    current_batch = ""

    def detect_batch_group(text: str):
        normalized = str(text).strip().lower().replace(" ", "")
        normalized_compact = re.sub(r"[^a-z0-9\u4e00-\u9fff]", "", normalized)
        if normalized in {"2", "02"} or normalized_compact in {"2", "02"}:
            return "batch2"
        if normalized in {"3", "03"} or normalized_compact in {"3", "03"}:
            return "batch3"
        if any(token in normalized for token in ["\u7b2c\u4e8c\u6279", "batch2", "batch_2", "batch-2", "2nd", "b2"]):
            return "batch2"
        if any(token in normalized for token in ["\u7b2c\u4e09\u6279", "batch3", "batch_3", "batch-3", "3rd", "b3"]):
            return "batch3"
        if re.fullmatch(r"batch[_-]?2", normalized_compact):
            return "batch2"
        if re.fullmatch(r"batch[_-]?3", normalized_compact):
            return "batch3"
        return None

    rules = []
    for row in rows[1:]:
        batch_cell = row[idx_batch] if idx_batch < len(row) else None
        list_cell = row[idx_list] if idx_list < len(row) else None
        type_cell = row[idx_type] if idx_type < len(row) else None
        sex_cell = row[idx_sex] if idx_sex < len(row) else None
        if batch_cell:
            current_batch = str(batch_cell).strip()
        if not current_batch or not list_cell or not type_cell or not sex_cell:
            continue
        sex = str(sex_cell).strip().upper()
        if sex not in {"XX", "XY"}:
            continue
        batch_group = detect_batch_group(current_batch)
        if batch_group is None:
            continue
        rules.append(
            {
                "batch": current_batch,
                "batch_group": batch_group,
                "sex": sex,
                "ids": parse_id_tokens(str(list_cell)),
            }
        )
    return rules


def key_for_batch2(sample_id: str):
    token = re.split(r"[_\-]", sample_id)[0]
    m = re.match(r"([A-Za-z])", token)
    return m.group(1).upper() if m else None


def key_for_batch3(sample_id: str):
    s = sample_id.replace("_", "-")
    m = re.search(r"-([0-9]+)-([0-9]+)(?:[._-].*)?$", s)
    if m:
        return str(int(m.group(1)))
    m = re.search(r"-([0-9]+)(?:[._-].*)?$", s)
    if m:
        return str(int(m.group(1)))
    return None


def build_groups_by_rules(rules, detected, key_getter):
    lookup = defaultdict(list)
    for sample_id in sorted(detected.keys()):
        key = key_getter(sample_id)
        if key is not None:
            lookup[str(key).upper()].append(sample_id)

    groups = {"XX": [], "XY": []}
    for rule in rules:
        for token in rule["ids"]:
            token_norm = str(token).upper()
            matched = sorted(lookup.get(token_norm, []))
            for sample_id in matched:
                if sample_id not in groups[rule["sex"]]:
                    groups[rule["sex"]].append(sample_id)
    groups["XX"] = sorted(groups["XX"])
    groups["XY"] = sorted(groups["XY"])
    return groups


def sort_group_items(items):
    unique_items = set(items)
    return sorted(unique_items, key=lambda x: (1, int(str(x))) if str(x).isdigit() else (0, str(x)))


def select_reference_groups(sample_info_xlsx: Path, batch2_dir: Path, batch3_dir: Path):
    rules = parse_reference_rules_from_excel(sample_info_xlsx)

    allowed_batch2_ids = {chr(x) for x in range(ord("A"), ord("H") + 1)}
    rules_batch2 = []
    for rule in rules:
        if rule.get("batch_group") != "batch2":
            continue
        if rule.get("sex") != "XY":
            continue
        filtered_ids = [token for token in rule.get("ids", []) if str(token).upper() in allowed_batch2_ids]
        if not filtered_ids:
            continue
        new_rule = dict(rule)
        new_rule["ids"] = filtered_ids
        rules_batch2.append(new_rule)
    rules_batch3 = [r for r in rules if r.get("batch_group") == "batch3"]

    samples_batch2 = detect_samples(batch2_dir)
    samples_batch3 = detect_samples(batch3_dir)

    groups2 = build_groups_by_rules(rules_batch2, samples_batch2, key_for_batch2)
    allowed_batch2_heads = {chr(x) for x in range(ord("A"), ord("H") + 1)}
    groups2["XX"] = []
    groups2["XY"] = sorted({
        key_for_batch2(sample_id)
        for sample_id in groups2.get("XY", [])
        if key_for_batch2(sample_id) in allowed_batch2_heads
    })
    groups3 = build_groups_by_rules(rules_batch3, samples_batch3, key_for_batch3)
    groups3["XX"] = sorted({key_for_batch3(sample_id) for sample_id in groups3.get("XX", []) if key_for_batch3(sample_id)}, key=int)
    groups3["XY"] = sorted({key_for_batch3(sample_id) for sample_id in groups3.get("XY", []) if key_for_batch3(sample_id)}, key=int)

    merged = {"XX": [], "XY": []}
    for sex in ["XX", "XY"]:
        merged[sex] = sort_group_items(groups2.get(sex, []) + groups3.get(sex, []))
    return merged


def merge_samples_with_batch(sample_dict, batch_label):
    merged = {}
    for sample_id, pair in sample_dict.items():
        key = sample_id
        if key in merged:
            key = f"{sample_id}_{batch_label}"
        merged[key] = {"R1": str(Path(pair["R1"]).resolve()), "R2": str(Path(pair["R2"]).resolve())}
    return merged


def remap_batch2_samples_to_letter_id(sample_dict):
    remapped = {}
    for sample_id, pair in sample_dict.items():
        letter_id = key_for_batch2(sample_id)
        key = letter_id if letter_id is not None else sample_id
        if key in remapped:
            raise ValueError(f"Duplicate batch2 sample letter id detected: {key}")
        remapped[key] = pair
    return remapped


def remap_batch3_samples_to_numeric_id(sample_dict):
    remapped = {}
    for sample_id, pair in sample_dict.items():
        numeric_id = key_for_batch3(sample_id)
        key = numeric_id if numeric_id is not None else sample_id
        if key in remapped:
            raise ValueError(f"Duplicate batch3 sample numeric id detected: {key}")
        remapped[key] = pair
    return remapped


def resolve_sample_info_xlsx(path_value: str, project_root: Path | None = None):
    candidate = Path(path_value)
    if candidate.exists():
        return candidate
    fallbacks = [Path("sample_info/CNV-seq.xlsx"), Path("rules/sample_info/CNV-seq.xlsx")]
    if project_root is not None:
        fallbacks = [project_root / "sample_info/CNV-seq.xlsx", project_root / "rules/sample_info/CNV-seq.xlsx"] + fallbacks
    for fallback in fallbacks:
        if fallback.exists():
            return fallback
    return candidate


def build_reference_groups_into_config(
    project: str,
    config_path: Path,
    sample_info_xlsx: str,
    batch2_fq_dir: str,
    batch3_fq_dir: str,
):
    project_root = Path(project).resolve()
    config_file = config_path if config_path.is_absolute() else (project_root / config_path)
    config_found = config_file.exists()
    if not config_found:
        alt = project_root / "config.yaml"
        if alt.exists():
            config_file = alt
            config_found = True

    if config_found:
        with open(config_file, "r", encoding="utf-8") as handle:
            config = yaml.safe_load(handle) or {}
    else:
        config = load_template_config()
        config["core"]["project_path"] = str(project_root)

    samples_batch2 = remap_batch2_samples_to_letter_id(detect_samples(Path(batch2_fq_dir)))
    samples_batch3 = remap_batch3_samples_to_numeric_id(detect_samples(Path(batch3_fq_dir)))
    merged_samples = {}
    merged_samples.update(merge_samples_with_batch(samples_batch2, "B2"))
    for sample_id, pair in merge_samples_with_batch(samples_batch3, "B3").items():
        if sample_id in merged_samples:
            sample_id = f"{sample_id}_B3"
        merged_samples[sample_id] = pair
    config["samples"] = merged_samples

    sample_info = resolve_sample_info_xlsx(sample_info_xlsx, project_root=project_root)
    groups = select_reference_groups(
        sample_info_xlsx=sample_info,
        batch2_dir=Path(batch2_fq_dir),
        batch3_dir=Path(batch3_fq_dir),
    )
    config.setdefault("build_reference", {})
    config["build_reference"]["enabled"] = True
    config["build_reference"]["mode"] = "excel"
    config["build_reference"]["sample_info_xlsx"] = str(sample_info.resolve())
    config["build_reference"]["rawdata_dirs"] = {
        "batch2": str(batch2_fq_dir),
        "batch3": str(batch3_fq_dir),
    }
    config["build_reference"]["groups"] = groups

    config_file.parent.mkdir(parents=True, exist_ok=True)
    with open(config_file, "w", encoding="utf-8") as handle:
        yaml.safe_dump(config, handle, sort_keys=False, allow_unicode=True)
    LOGGER.info("build_reference groups updated: XX=%d XY=%d", len(groups["XX"]), len(groups["XY"]))
    LOGGER.info("config updated: %s", config_file)


def init_config(
    project,
    fq_dir,
    output_config,
    template_snakefile=None,
    build_reference_mode="none",
    sample_info_xlsx="sample_info/CNV-seq.xlsx",
    batch2_fq_dir="/data/project/CNV/PGT-A/rawdata/lib_test/2026-02-05",
    batch3_fq_dir="/data/project/CNV/PGT-A/rawdata/lib_test/2026-03-03",
):
    project = Path(project).resolve()
    fq_dir = Path(fq_dir).resolve()
    config_path = Path(output_config).resolve()

    project.mkdir(parents=True, exist_ok=True)
    for sub in [
        "data/fastq",
        "fastp",
        "mapping",
        "wisecondorx/converted",
        "wisecondorx/reference",
        "wisecondorx/tuning",
        "reference/XX/prefilter",
        "reference/XX/tuning",
        "reference/XX/result",
        "reference/XY/prefilter",
        "reference/XY/tuning",
        "reference/XY/result",
        "wisecondorx/cnv",
        "wisecondorx/cnv/qc",
        "wisecondorx/cnv/predict",
        "logs/fastp",
        "logs/bwa",
        "logs/metadata",
        "logs/wisecondorx",
        "logs/cnv",
    ]:
        (project / sub).mkdir(parents=True, exist_ok=True)

    samples = detect_samples(fq_dir)
    for sample_id, pair in samples.items():
        for direction in ["R1", "R2"]:
            src = pair[direction].resolve()
            dst = project / "data/fastq" / src.name
            if dst.exists() or dst.is_symlink():
                dst.unlink()
            dst.symlink_to(src)
            samples[sample_id][direction] = str(dst.resolve())

    config = load_template_config()
    config["core"]["project_path"] = str(project)
    config["samples"] = samples

    if build_reference_mode == "excel":
        sample_info = resolve_sample_info_xlsx(sample_info_xlsx, project_root=project)
        groups = select_reference_groups(
            sample_info_xlsx=sample_info,
            batch2_dir=Path(batch2_fq_dir),
            batch3_dir=Path(batch3_fq_dir),
        )
        config["build_reference"] = {}
        config["build_reference"]["enabled"] = True
        config["build_reference"]["mode"] = "excel"
        config["build_reference"]["sample_info_xlsx"] = str(sample_info.resolve())
        config["build_reference"]["rawdata_dirs"] = {
            "batch2": str(batch2_fq_dir),
            "batch3": str(batch3_fq_dir),
        }
        config["build_reference"]["groups"] = groups
        LOGGER.info("reference groups from excel: XX=%d XY=%d", len(groups["XX"]), len(groups["XY"]))

    config_path.parent.mkdir(parents=True, exist_ok=True)
    with open(config_path, "w", encoding="utf-8") as handle:
        yaml.safe_dump(config, handle, sort_keys=False, allow_unicode=True)
    LOGGER.info("config written: %s", config_path)

    if template_snakefile:
        snakefile_src = Path(template_snakefile).resolve()
        snakefile_dst = project / "Snakefile"
        if snakefile_src.exists():
            shutil.copy(snakefile_src, snakefile_dst)
            LOGGER.info("Snakefile copied to %s", snakefile_dst)
        else:
            LOGGER.warning("Snakefile not found at %s", snakefile_src)


if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description="Init PGT-A config or update build-reference groups from excel"
    )
    parser.add_argument(
        "--action",
        choices=["init_config", "build_reference_groups"],
        default="init_config",
        help="Run mode: init config or update build-reference groups only",
    )
    parser.add_argument("--project", required=False, help="Project path (analysis directory)")
    parser.add_argument("--fq_dir", required=False, help="Path to input fastq directory")
    parser.add_argument("--output_config", required=True, help="Path to target config.yaml")
    parser.add_argument(
        "--build_reference_mode",
        default="none",
        choices=["none", "excel"],
        help="Reference group selection mode",
    )
    parser.add_argument(
        "--sample_info_xlsx",
        default="sample_info/CNV-seq.xlsx",
        help="Excel with batch, sample list and sex",
    )
    parser.add_argument(
        "--batch2_fq_dir",
        default="/data/project/CNV/PGT-A/rawdata/lib_test/2026-02-05",
        help="Batch2 FASTQ dir (second batch, letter IDs)",
    )
    parser.add_argument(
        "--batch3_fq_dir",
        default="/data/project/CNV/PGT-A/rawdata/lib_test/2026-03-03",
        help="Batch3 FASTQ dir (third batch, JZ2604xxx-ID-ID)",
    )
    parser.add_argument(
        "--template_snakefile",
        required=False,
        help="Optional path to template Snakefile",
    )
    parser.add_argument("--log", default="", help="Optional log file path")
    args = parser.parse_args()
    LOGGER = setup_logger("init_wes_project", args.log or None)
    if args.action == "init_config":
        if not args.project or not args.fq_dir:
            raise ValueError("--project and --fq_dir are required for --action init_config")
        init_config(
            project=args.project,
            fq_dir=args.fq_dir,
            output_config=args.output_config,
            template_snakefile=args.template_snakefile,
            build_reference_mode=args.build_reference_mode,
            sample_info_xlsx=args.sample_info_xlsx,
            batch2_fq_dir=args.batch2_fq_dir,
            batch3_fq_dir=args.batch3_fq_dir,
        )
    else:
        if not args.project:
            raise ValueError("--project is required for --action build_reference_groups")
        build_reference_groups_into_config(
            project=args.project,
            config_path=Path(args.output_config).resolve(),
            sample_info_xlsx=args.sample_info_xlsx,
            batch2_fq_dir=args.batch2_fq_dir,
            batch3_fq_dir=args.batch3_fq_dir,
        )
