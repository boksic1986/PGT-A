#!/biosoftware/miniconda/envs/snakemake_env/bin/python
# -*- coding: utf-8 -*-
# Author: jiucheng
# Updated: 2026/03/09
import argparse
import datetime as dt
import logging
import re
import shutil
from collections import defaultdict
from pathlib import Path

from pgta.core.config import load_yaml_file, resolve_existing_path, write_yaml_file
from pgta.core.logging import setup_logger

LOGGER = logging.getLogger("init_pgta_project")


def normalize_header_name(value):
    return re.sub(r"\s+", "", str(value).strip().lower()) if value is not None else ""


def resolve_column_indices(header, aliases_by_key):
    normalized_header = [normalize_header_name(item) for item in header]
    col_map = {name: idx for idx, name in enumerate(normalized_header) if name}

    resolved = {}
    missing = []
    for key, aliases in aliases_by_key.items():
        idx = None
        for alias in aliases:
            alias_norm = normalize_header_name(alias)
            if alias_norm in col_map:
                idx = col_map[alias_norm]
                break
        if idx is None:
            missing.append(key)
        else:
            resolved[key] = idx
    if missing:
        raise ValueError(
            "Excel is missing required logical columns: "
            + ",".join(missing)
            + f". Header={header}"
        )
    return resolved


def normalize_sex_value(value):
    text = str(value).strip().upper().replace(" ", "")
    if text in {"XX", "XY"}:
        return text
    if "XY" in text:
        return "XY"
    if "XX" in text:
        return "XX"
    return ""


def detect_samples(fq_dir: Path) -> dict:
    fq_dir = Path(fq_dir)
    fq_files = (
        list(fq_dir.rglob("*.fastq.gz"))
        + list(fq_dir.rglob("*.fq.gz"))
        + list(fq_dir.rglob("*.fastq"))
        + list(fq_dir.rglob("*.fq"))
    )

    temp_dict = defaultdict(list)
    for fq in fq_files:
        name = fq.name
        base = re.sub(r"(?i)\.(fastq|fq)(\.gz)?$", "", name)
        match = re.search(r"(?i)(?:^|[._-])R?([12])(?:[._-]\d{3})?$", base)
        if not match:
            continue

        direction = "R1" if match.group(1) == "1" else "R2"
        sample_id = base[: match.start()].rstrip("._-")
        if not sample_id:
            continue
        temp_dict[sample_id].append((direction, fq))

    sample_dict = {}
    for sample_id, items in temp_dict.items():
        r1_list = sorted([fq for d, fq in items if d == "R1"])
        r2_list = sorted([fq for d, fq in items if d == "R2"])
        count = min(len(r1_list), len(r2_list))
        for i in range(count):
            key = f"{sample_id}_{i + 1}" if count > 1 else sample_id
            sample_dict[key] = {"R1": r1_list[i], "R2": r2_list[i]}
    return sample_dict


def load_template_config() -> dict:
    return {
        "core": {
            "project_path": "results",
            "reference_genome": "/data/Database/index/hg19/hg19.fa",
            "chromosome_list": [
                "chr1", "chr2", "chr3", "chr4", "chr5", "chr6", "chr7", "chr8", "chr9", "chr10",
                "chr11", "chr12", "chr13", "chr14", "chr15", "chr16", "chr17", "chr18", "chr19", "chr20",
                "chr21", "chr22", "chrX", "chrY",
            ],
            "wisecondorx": {
                "binsize": 100000,
                "reference_output": "reference/result/ref_best.npz",
                "reference_output_by_sex": {
                    "XX": "reference/XX/result/ref_xx_best.npz",
                    "XY": "reference/XY/result/ref_xy_best.npz",
                },
                "gender_reference_output": "reference/gender/result/ref_gender_best.npz",
                "common_reference_binsize_output": "reference/gender/common_best_binsize.txt",
                "reference_model_root": "reference",
                "use_chr_prefix": True,
                "tuning": {
                    "enable": True,
                    "bin_sizes": [100000, 200000, 500000, 750000, 1000000, 1500000],
                    "pca": {"min_components": 2, "max_components": 20, "min_explained_variance": 0.0},
                    "qc": {
                        "min_reference_samples": 8,
                        "max_outlier_fraction": 0.25,
                        "min_reads_per_sample": 3000000,
                        "min_corr_to_median": 0.9,
                        "max_reconstruction_error_z": 3.5,
                        "max_noise_mad_z": 3.5,
                    },
                },
                "reference_prefilter": {
                    "binsize": 100000,
                    "max_iterations": 3,
                },
                "cnv": {
                    "enable": True,
                    "output_dir": "wisecondorx/cnv",
                    "gender_dir": "wisecondorx/cnv/gender",
                    "qc_dir": "wisecondorx/cnv/qc",
                    "predict_dir": "wisecondorx/cnv/predict",
                    "convert_binsize": 100000,
                    "zscore": 5,
                    "alpha": 0.001,
                    "maskrepeats": 5,
                    "minrefbins": 150,
                    "qc": {"min_total_counts": 1000000, "min_nonzero_fraction": 0.4, "max_mad_log1p": 1.2},
                },
            },
        },
        "pipeline": {"mode": "build_ref", "targets": ["mapping", "metadata", "reference_qc", "reference"]},
        "build_reference": {
            "reference_sets": {
                "pass_only": {"decisions": ["PASS"], "publish_as_default": True},
                "pass_warn": {"decisions": ["PASS", "WARN"], "publish_as_default": False},
            }
        },
        "biosoft": {
            "fastp": "/biosoftware/bin/fastp",
            "bwa": "/biosoftware/bin/bwa",
            "samtools": "/biosoftware/bin/samtools",
            "WisecondorX": "/biosoftware/miniconda/envs/wise_env/bin/WisecondorX",
            "python": "/biosoftware/miniconda/envs/snakemake_env/bin/python",
        },
        "samples": {},
    }


def parse_id_tokens(text: str):
    normalized = (
        str(text)
        .strip()
        .replace("\u2014", "-")
        .replace("\u2013", "-")
        .replace("\uff0c", ",")
        .replace(" ", "")
    )
    tokens = []
    for item in normalized.split(","):
        if not item:
            continue
        if "-" in item:
            left, right = item.split("-", 1)
            left_prefix_match = re.fullmatch(r"([A-Za-z]+)([0-9]+)", left)
            right_prefix_match = re.fullmatch(r"([A-Za-z]+)([0-9]+)", right)
            right_numeric_match = re.fullmatch(r"([0-9]+)", right)
            if left.isdigit() and right.isdigit():
                a, b = int(left), int(right)
                if b < a:
                    a, b = b, a
                tokens.extend([str(x) for x in range(a, b + 1)])
                continue
            if re.fullmatch(r"[A-Za-z]", left) and re.fullmatch(r"[A-Za-z]", right):
                a, b = ord(left.upper()), ord(right.upper())
                if b < a:
                    a, b = b, a
                tokens.extend([chr(x) for x in range(a, b + 1)])
                continue
            if left_prefix_match and right_prefix_match and left_prefix_match.group(1).upper() == right_prefix_match.group(1).upper():
                prefix = left_prefix_match.group(1).upper()
                a, b = int(left_prefix_match.group(2)), int(right_prefix_match.group(2))
                if b < a:
                    a, b = b, a
                tokens.extend([f"{prefix}{x}" for x in range(a, b + 1)])
                continue
            if left_prefix_match and right_numeric_match:
                prefix = left_prefix_match.group(1).upper()
                a, b = int(left_prefix_match.group(2)), int(right_numeric_match.group(1))
                if b < a:
                    a, b = b, a
                tokens.extend([f"{prefix}{x}" for x in range(a, b + 1)])
                continue
        tokens.append(item.upper() if re.fullmatch(r"[A-Za-z]+", item) else item)
    return tokens


def normalize_sample_token(token: str):
    text = str(token).strip().replace("_", "").replace(" ", "")
    return text.upper()


def extract_terminal_repeated_token(name: str):
    normalized = str(name).strip()
    match = re.search(r"[-_]([A-Za-z]*\d+|[A-Za-z]+)[-_]([A-Za-z]*\d+|[A-Za-z]+)$", normalized)
    if not match:
        return None
    left = normalize_sample_token(match.group(1))
    right = normalize_sample_token(match.group(2))
    if left != right:
        return None
    return left


def detect_selected_samples_from_rawdata(rawdata_root: Path, selected_ids):
    rawdata_root = Path(rawdata_root)
    if not rawdata_root.exists():
        raise FileNotFoundError(f"Rawdata root not found: {rawdata_root}")

    requested_ids = {normalize_sample_token(item) for item in selected_ids}
    if not requested_ids:
        raise ValueError("No selected sample IDs provided.")

    found = {}
    found_sources = {}
    found_sort_keys = {}
    candidate_dirs = sorted(path for path in rawdata_root.rglob("*") if path.is_dir())
    for sample_dir in candidate_dirs:
        sample_token = extract_terminal_repeated_token(sample_dir.name)
        if sample_token is None or sample_token not in requested_ids:
            continue
        detected = detect_samples(sample_dir)
        if not detected:
            continue
        if len(detected) != 1:
            raise ValueError(
                f"Expected exactly one FASTQ sample pair under {sample_dir}, found {len(detected)}: {','.join(sorted(detected))}"
            )
        pair = next(iter(detected.values()))
        sort_key = build_sample_dir_sort_key(sample_dir, pair)
        if sample_token in found:
            previous_source = found_sources[sample_token]
            previous_sort_key = found_sort_keys[sample_token]
            if sort_key > previous_sort_key:
                LOGGER.warning(
                    "duplicate selected sample ID detected: %s; keep newer dir=%s replace older dir=%s",
                    sample_token,
                    sample_dir,
                    previous_source,
                )
            else:
                LOGGER.warning(
                    "duplicate selected sample ID detected: %s; keep newer dir=%s ignore older dir=%s",
                    sample_token,
                    previous_source,
                    sample_dir,
                )
                continue
        found[sample_token] = {"R1": Path(pair["R1"]).resolve(), "R2": Path(pair["R2"]).resolve()}
        found_sources[sample_token] = str(sample_dir.resolve())
        found_sort_keys[sample_token] = sort_key

    missing = sorted(requested_ids - set(found))
    if missing:
        raise ValueError(
            "Selected sample IDs were not found in rawdata root: "
            + ",".join(missing)
            + f" (rawdata_root={rawdata_root})"
        )
    return found


def extract_datetimes_from_text(text: str):
    text = str(text)
    values = []

    for match in re.finditer(r"(20\d{2})[-_]?(\d{2})[-_]?(\d{2})[-_]?(\d{2})(\d{2})(\d{2})", text):
        year, month, day, hour, minute, second = map(int, match.groups())
        try:
            values.append(dt.datetime(year, month, day, hour, minute, second))
        except ValueError:
            continue

    for match in re.finditer(r"(20\d{2})[-_]?(\d{2})[-_]?(\d{2})[-_]?(\d{2})(\d{2})", text):
        year, month, day, hour, minute = map(int, match.groups())
        try:
            values.append(dt.datetime(year, month, day, hour, minute, 0))
        except ValueError:
            continue

    return values


def extract_latest_datetime_from_path(path_value: Path):
    candidates = []
    for item in [Path(path_value)] + list(Path(path_value).parents):
        candidates.extend(extract_datetimes_from_text(item.name))
    if candidates:
        return max(candidates)
    return None


def build_sample_dir_sort_key(sample_dir: Path, pair):
    latest_dt = extract_latest_datetime_from_path(sample_dir)
    if latest_dt is None:
        timestamp_value = 0.0
    else:
        timestamp_value = latest_dt.timestamp()
    file_mtime = max(Path(pair["R1"]).stat().st_mtime, Path(pair["R2"]).stat().st_mtime)
    return (timestamp_value, file_mtime, str(sample_dir.resolve()))


def load_sample_sex_map_from_excel(sample_info_xlsx: Path):
    import openpyxl

    wb = openpyxl.load_workbook(sample_info_xlsx, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise ValueError(f"Empty excel file: {sample_info_xlsx}")

    header = [str(x).strip() if x is not None else "" for x in rows[0]]
    column_indices = resolve_column_indices(
        header,
        {
            "sample_ids": [
                "sample_ids",
                "样本清单",
                "cnv-seq测序数据编号",
                "cnvseq测序数据编号",
                "sample id",
                "sample_id",
            ],
            "sex": ["sex", "性别"],
        },
    )

    idx_list = column_indices["sample_ids"]
    idx_sex = column_indices["sex"]
    sex_map = {}
    for row in rows[1:]:
        list_cell = row[idx_list] if idx_list < len(row) else None
        sex_cell = row[idx_sex] if idx_sex < len(row) else None
        if not list_cell or not sex_cell:
            continue
        sex = normalize_sex_value(sex_cell)
        if sex not in {"XX", "XY"}:
            continue
        for token in parse_id_tokens(str(list_cell)):
            token_norm = normalize_sample_token(token)
            previous = sex_map.get(token_norm)
            if previous and previous != sex:
                raise ValueError(f"Conflicting sex assignment for sample token {token_norm}: {previous} vs {sex}")
            sex_map[token_norm] = sex
    return sex_map


def build_reference_groups_for_selected_ids(selected_ids, sample_info_xlsx: Path):
    sex_map = load_sample_sex_map_from_excel(sample_info_xlsx)
    groups = {"XX": [], "XY": []}
    missing = []
    for sample_id in sorted({normalize_sample_token(item) for item in selected_ids}):
        sex = sex_map.get(sample_id)
        if sex is None:
            missing.append(sample_id)
            continue
        groups[sex].append(sample_id)
    if missing:
        raise ValueError(
            "Selected baseline sample IDs are missing from sample info excel: "
            + ",".join(missing)
            + f" ({sample_info_xlsx})"
        )
    return groups


def parse_reference_rules_from_excel(sample_info_xlsx: Path):
    import openpyxl

    wb = openpyxl.load_workbook(sample_info_xlsx, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise ValueError(f"Empty excel file: {sample_info_xlsx}")

    header = [str(x).strip() if x is not None else "" for x in rows[0]]
    column_indices = resolve_column_indices(
        header,
        {
            "batch": ["batch", "批次"],
            "sample_ids": ["sample_ids", "样本清单", "sampleid", "sample_id"],
            "sample_type": ["sample_type", "样本类型"],
            "sex": ["sex", "性别"],
        },
    )

    idx_batch = column_indices["batch"]
    idx_list = column_indices["sample_ids"]
    idx_type = column_indices["sample_type"]
    idx_sex = column_indices["sex"]

    current_batch = ""

    def detect_batch_group(text: str):
        normalized = str(text).strip().lower().replace(" ", "")
        normalized_compact = re.sub(r"[^a-z0-9\u4e00-\u9fff]", "", normalized)
        if normalized in {"2", "02"} or normalized_compact in {"2", "02"}:
            return "batch2"
        if normalized in {"3", "03"} or normalized_compact in {"3", "03"}:
            return "batch3"
        if any(token in normalized for token in ["\u7b2c\u4e8c\u6279", "batch2", "batch_2", "batch-2", "2nd", "b2"]):
            return "batch2"
        if any(token in normalized for token in ["\u7b2c\u4e09\u6279", "batch3", "batch_3", "batch-3", "3rd", "b3"]):
            return "batch3"
        if re.fullmatch(r"batch[_-]?2", normalized_compact):
            return "batch2"
        if re.fullmatch(r"batch[_-]?3", normalized_compact):
            return "batch3"
        return None

    rules = []
    for row in rows[1:]:
        batch_cell = row[idx_batch] if idx_batch < len(row) else None
        list_cell = row[idx_list] if idx_list < len(row) else None
        type_cell = row[idx_type] if idx_type < len(row) else None
        sex_cell = row[idx_sex] if idx_sex < len(row) else None
        if batch_cell:
            current_batch = str(batch_cell).strip()
        if not current_batch or not list_cell or not type_cell or not sex_cell:
            continue
        sex = normalize_sex_value(sex_cell)
        if sex not in {"XX", "XY"}:
            continue
        batch_group = detect_batch_group(current_batch)
        if batch_group is None:
            continue
        rules.append(
            {
                "batch": current_batch,
                "batch_group": batch_group,
                "sex": sex,
                "ids": parse_id_tokens(str(list_cell)),
            }
        )
    return rules


def key_for_batch2(sample_id: str):
    token = re.split(r"[_\-]", sample_id)[0]
    m = re.match(r"([A-Za-z])", token)
    return m.group(1).upper() if m else None


def key_for_batch3(sample_id: str):
    s = sample_id.replace("_", "-")
    m = re.search(r"-([0-9]+)-([0-9]+)(?:[._-].*)?$", s)
    if m:
        return str(int(m.group(1)))
    m = re.search(r"-([0-9]+)(?:[._-].*)?$", s)
    if m:
        return str(int(m.group(1)))
    return None


def build_groups_by_rules(rules, detected, key_getter):
    lookup = defaultdict(list)
    for sample_id in sorted(detected.keys()):
        key = key_getter(sample_id)
        if key is not None:
            lookup[str(key).upper()].append(sample_id)

    groups = {"XX": [], "XY": []}
    for rule in rules:
        for token in rule["ids"]:
            token_norm = str(token).upper()
            matched = sorted(lookup.get(token_norm, []))
            for sample_id in matched:
                if sample_id not in groups[rule["sex"]]:
                    groups[rule["sex"]].append(sample_id)
    groups["XX"] = sorted(groups["XX"])
    groups["XY"] = sorted(groups["XY"])
    return groups


def sort_group_items(items):
    unique_items = set(items)
    return sorted(unique_items, key=lambda x: (1, int(str(x))) if str(x).isdigit() else (0, str(x)))


def select_reference_groups(sample_info_xlsx: Path, batch2_dir: Path, batch3_dir: Path):
    rules = parse_reference_rules_from_excel(sample_info_xlsx)

    allowed_batch2_ids = {chr(x) for x in range(ord("A"), ord("H") + 1)}
    rules_batch2 = []
    for rule in rules:
        if rule.get("batch_group") != "batch2":
            continue
        if rule.get("sex") != "XY":
            continue
        filtered_ids = [token for token in rule.get("ids", []) if str(token).upper() in allowed_batch2_ids]
        if not filtered_ids:
            continue
        new_rule = dict(rule)
        new_rule["ids"] = filtered_ids
        rules_batch2.append(new_rule)
    rules_batch3 = [r for r in rules if r.get("batch_group") == "batch3"]

    samples_batch2 = detect_samples(batch2_dir)
    samples_batch3 = detect_samples(batch3_dir)

    groups2 = build_groups_by_rules(rules_batch2, samples_batch2, key_for_batch2)
    allowed_batch2_heads = {chr(x) for x in range(ord("A"), ord("H") + 1)}
    groups2["XX"] = []
    groups2["XY"] = sorted({
        key_for_batch2(sample_id)
        for sample_id in groups2.get("XY", [])
        if key_for_batch2(sample_id) in allowed_batch2_heads
    })
    groups3 = build_groups_by_rules(rules_batch3, samples_batch3, key_for_batch3)
    groups3["XX"] = sorted({key_for_batch3(sample_id) for sample_id in groups3.get("XX", []) if key_for_batch3(sample_id)}, key=int)
    groups3["XY"] = sorted({key_for_batch3(sample_id) for sample_id in groups3.get("XY", []) if key_for_batch3(sample_id)}, key=int)

    merged = {"XX": [], "XY": []}
    for sex in ["XX", "XY"]:
        merged[sex] = sort_group_items(groups2.get(sex, []) + groups3.get(sex, []))
    return merged


def merge_samples_with_batch(sample_dict, batch_label):
    merged = {}
    for sample_id, pair in sample_dict.items():
        key = sample_id
        if key in merged:
            key = f"{sample_id}_{batch_label}"
        merged[key] = {"R1": str(Path(pair["R1"]).resolve()), "R2": str(Path(pair["R2"]).resolve())}
    return merged


def remap_batch2_samples_to_letter_id(sample_dict):
    remapped = {}
    for sample_id, pair in sample_dict.items():
        letter_id = key_for_batch2(sample_id)
        key = letter_id if letter_id is not None else sample_id
        if key in remapped:
            raise ValueError(f"Duplicate batch2 sample letter id detected: {key}")
        remapped[key] = pair
    return remapped


def remap_batch3_samples_to_numeric_id(sample_dict):
    remapped = {}
    for sample_id, pair in sample_dict.items():
        numeric_id = key_for_batch3(sample_id)
        key = numeric_id if numeric_id is not None else sample_id
        if key in remapped:
            raise ValueError(f"Duplicate batch3 sample numeric id detected: {key}")
        remapped[key] = pair
    return remapped


def resolve_sample_info_xlsx(path_value: str, project_root: Path | None = None):
    return resolve_existing_path(
        path_value=path_value,
        label="sample_info_xlsx",
        project_root=project_root,
        expect_dir=False,
    )


def detect_samples_or_fail(fq_dir: Path, label: str):
    samples = detect_samples(fq_dir)
    if not samples:
        raise ValueError(f"No paired FASTQ samples detected under --{label}: {fq_dir}")
    return samples


def build_reference_groups_into_config(
    project: str,
    config_path: Path,
    sample_info_xlsx: str,
    batch2_fq_dir: str,
    batch3_fq_dir: str,
):
    project_root = Path(project).resolve()
    config_file = config_path if config_path.is_absolute() else (project_root / config_path)
    config_found = config_file.exists()
    if not config_found:
        for alt_name in ("build_samples.yaml", "config.yaml"):
            alt = project_root / alt_name
            if alt.exists():
                config_file = alt
                config_found = True
                break

    if config_found:
        config = load_yaml_file(config_file)
    else:
        config = load_template_config()
        config["core"]["project_path"] = str(project_root)

    batch2_dir = resolve_existing_path(batch2_fq_dir, "batch2_fq_dir", project_root=project_root, expect_dir=True)
    batch3_dir = resolve_existing_path(batch3_fq_dir, "batch3_fq_dir", project_root=project_root, expect_dir=True)
    samples_batch2 = remap_batch2_samples_to_letter_id(detect_samples_or_fail(batch2_dir, "batch2_fq_dir"))
    samples_batch3 = remap_batch3_samples_to_numeric_id(detect_samples_or_fail(batch3_dir, "batch3_fq_dir"))
    merged_samples = {}
    merged_samples.update(merge_samples_with_batch(samples_batch2, "B2"))
    for sample_id, pair in merge_samples_with_batch(samples_batch3, "B3").items():
        if sample_id in merged_samples:
            sample_id = f"{sample_id}_B3"
        merged_samples[sample_id] = pair
    config["samples"] = merged_samples

    sample_info = resolve_sample_info_xlsx(sample_info_xlsx, project_root=project_root)
    groups = select_reference_groups(
        sample_info_xlsx=sample_info,
        batch2_dir=batch2_dir,
        batch3_dir=batch3_dir,
    )
    config.setdefault("build_reference", {})
    config["build_reference"]["enabled"] = True
    config["build_reference"]["mode"] = "excel"
    config["build_reference"]["reference_sets"] = {
        "pass_only": {"decisions": ["PASS"], "publish_as_default": True},
        "pass_warn": {"decisions": ["PASS", "WARN"], "publish_as_default": False},
    }
    config["build_reference"]["sample_info_xlsx"] = str(sample_info.resolve())
    config["build_reference"]["rawdata_dirs"] = {
        "batch2": str(batch2_dir),
        "batch3": str(batch3_dir),
    }
    config["build_reference"]["groups"] = groups
    config.setdefault("pipeline", {})
    config["pipeline"]["mode"] = "build_ref"
    config["pipeline"]["targets"] = ["mapping", "metadata", "reference_qc", "reference"]

    write_yaml_file(config_file, config)
    LOGGER.info("build_reference groups updated: XX=%d XY=%d", len(groups["XX"]), len(groups["XY"]))
    LOGGER.info("config updated: %s", config_file)


def init_config(
    project,
    fq_dir,
    output_config,
    template_snakefile=None,
    config_kind="auto",
    build_reference_mode="none",
    sample_info_xlsx="",
    batch2_fq_dir="",
    batch3_fq_dir="",
    rawdata_root="",
    sample_ids="",
):
    project = Path(project).resolve()
    config_path = Path(output_config).resolve()
    rawdata_root = resolve_existing_path(rawdata_root, "rawdata_root", expect_dir=True) if rawdata_root else None
    sample_ids_list = parse_id_tokens(sample_ids) if sample_ids else []
    fq_dir_path = resolve_existing_path(fq_dir, "fq_dir", expect_dir=True) if fq_dir else None

    project.mkdir(parents=True, exist_ok=True)
    for sub in [
        "data/fastq",
        "fastp",
        "mapping",
        "wisecondorx/converted",
        "wisecondorx/tuning",
        "reference/prefilter",
        "reference/tuning",
        "reference/result",
        "reference/XX/prefilter",
        "reference/XX/tuning",
        "reference/XX/result",
        "reference/XY/prefilter",
        "reference/XY/tuning",
        "reference/XY/result",
        "reference/gender/result",
        "reference/cohorts/pass_only",
        "reference/cohorts/pass_warn",
        "wisecondorx/cnv",
        "wisecondorx/cnv/gender",
        "wisecondorx/cnv/qc",
        "wisecondorx/cnv/predict",
        "logs/fastp",
        "logs/bwa",
        "logs/metadata",
        "logs/wisecondorx",
        "logs/cnv",
    ]:
        (project / sub).mkdir(parents=True, exist_ok=True)

    if rawdata_root is not None:
        if not sample_ids_list:
            raise ValueError("--sample_ids is required when --rawdata_root is provided")
        samples = detect_selected_samples_from_rawdata(rawdata_root, sample_ids_list)
        if not samples:
            raise ValueError(f"No samples matched --sample_ids under --rawdata_root: {rawdata_root}")
    elif fq_dir_path is not None:
        samples = detect_samples_or_fail(fq_dir_path, "fq_dir")
    else:
        raise ValueError("Either --fq_dir or --rawdata_root with --sample_ids is required.")

    for sample_id, pair in samples.items():
        for direction in ["R1", "R2"]:
            src = pair[direction].resolve()
            dst = project / "data/fastq" / src.name
            if dst.exists() or dst.is_symlink():
                dst.unlink()
            dst.symlink_to(src)
            samples[sample_id][direction] = str(dst.resolve())

    config = load_template_config()
    config["core"]["project_path"] = str(project)
    config["samples"] = samples

    if config_kind == "auto":
        config_kind = "build_ref" if build_reference_mode == "excel" else "predict_samples"
    if config_kind not in {"build_ref", "predict_samples"}:
        raise ValueError("config_kind must be one of: auto, build_ref, predict_samples")

    if build_reference_mode == "excel":
        batch2_dir = resolve_existing_path(batch2_fq_dir, "batch2_fq_dir", project_root=project, expect_dir=True)
        batch3_dir = resolve_existing_path(batch3_fq_dir, "batch3_fq_dir", project_root=project, expect_dir=True)
        sample_info = resolve_sample_info_xlsx(sample_info_xlsx, project_root=project)
        groups = select_reference_groups(
            sample_info_xlsx=sample_info,
            batch2_dir=batch2_dir,
            batch3_dir=batch3_dir,
        )
        config["build_reference"] = {}
        config["build_reference"]["enabled"] = True
        config["build_reference"]["mode"] = "excel"
        config["build_reference"]["reference_sets"] = {
            "pass_only": {"decisions": ["PASS"], "publish_as_default": True},
            "pass_warn": {"decisions": ["PASS", "WARN"], "publish_as_default": False},
        }
        config["build_reference"]["sample_info_xlsx"] = str(sample_info.resolve())
        config["build_reference"]["rawdata_dirs"] = {
            "batch2": str(batch2_dir),
            "batch3": str(batch3_dir),
        }
        config["build_reference"]["groups"] = groups
        LOGGER.info("reference groups from excel: XX=%d XY=%d", len(groups["XX"]), len(groups["XY"]))
    elif config_kind == "build_ref":
        groups = {"XX": [], "XY": []}
        mode = "manual"
        if rawdata_root is not None:
            sample_info = resolve_sample_info_xlsx(sample_info_xlsx, project_root=project)
            groups = build_reference_groups_for_selected_ids(samples.keys(), sample_info)
            mode = "selected_ids"
            LOGGER.info("reference groups from selected IDs: XX=%d XY=%d", len(groups["XX"]), len(groups["XY"]))
        config["build_reference"] = {
            "enabled": True,
            "mode": mode,
            "groups": groups,
            "reference_sets": {
                "pass_only": {"decisions": ["PASS"], "publish_as_default": True},
                "pass_warn": {"decisions": ["PASS", "WARN"], "publish_as_default": False},
            },
        }
        if rawdata_root is not None:
            config["build_reference"]["sample_info_xlsx"] = str(sample_info.resolve())
            config["build_reference"]["rawdata_root"] = str(rawdata_root)
            config["build_reference"]["selected_ids"] = sorted(samples.keys())

    if config_kind == "build_ref":
        config["pipeline"]["mode"] = "build_ref"
        config["pipeline"]["targets"] = ["mapping", "metadata", "reference_qc", "reference"]
    else:
        config.pop("build_reference", None)
        config["pipeline"]["mode"] = "predict"
        config["pipeline"]["targets"] = ["mapping", "metadata", "cnv_qc", "cnv"]

    write_yaml_file(config_path, config)
    LOGGER.info("config written: %s", config_path)

    if template_snakefile:
        snakefile_src = Path(template_snakefile).resolve()
        snakefile_dst = project / "Snakefile"
        if snakefile_src.exists():
            shutil.copy(snakefile_src, snakefile_dst)
            LOGGER.info("Snakefile copied to %s", snakefile_dst)
        else:
            LOGGER.warning("Snakefile not found at %s", snakefile_src)


def main():
    parser = argparse.ArgumentParser(
        description="Init PGT-A config or update build-reference groups from excel"
    )
    parser.add_argument(
        "--action",
        choices=["init_config", "build_reference_groups"],
        default="init_config",
        help="Run mode: init config or update build-reference groups only",
    )
    parser.add_argument("--project", required=False, help="Project path (analysis directory)")
    parser.add_argument("--fq_dir", required=False, help="Path to input fastq directory")
    parser.add_argument(
        "--rawdata_root",
        required=False,
        default="",
        help="Root directory to recursively scan for sample subdirectories like Sample_XXX-E1-E1",
    )
    parser.add_argument(
        "--sample_ids",
        required=False,
        default="",
        help="Comma-separated sample IDs or ranges, e.g. A-H,1-10,12,13,E1-E16 or Y1-Y8",
    )
    parser.add_argument("--output_config", required=True, help="Path to target config.yaml")
    parser.add_argument(
        "--build_reference_mode",
        default="none",
        choices=["none", "excel"],
        help="Reference group selection mode",
    )
    parser.add_argument(
        "--config_kind",
        default="auto",
        choices=["auto", "build_ref", "predict_samples"],
        help="Config type to generate. auto => build_ref when build_reference_mode=excel, otherwise predict_samples.",
    )
    parser.add_argument(
        "--sample_info_xlsx",
        default="",
        help="Excel with batch, sample list and sex. Required for excel-based build-reference grouping.",
    )
    parser.add_argument(
        "--batch2_fq_dir",
        default="",
        help="Batch2 FASTQ dir (second batch, letter IDs). Required for excel-based build-reference grouping.",
    )
    parser.add_argument(
        "--batch3_fq_dir",
        default="",
        help="Batch3 FASTQ dir (third batch, JZ2604xxx-ID-ID). Required for excel-based build-reference grouping.",
    )
    parser.add_argument(
        "--template_snakefile",
        required=False,
        help="Optional path to template Snakefile",
    )
    parser.add_argument("--log", default="", help="Optional log file path")
    args = parser.parse_args()
    LOGGER = setup_logger("init_pgta_project", args.log or None)
    if args.action == "init_config":
        if not args.project:
            raise ValueError("--project is required for --action init_config")
        if not args.fq_dir and not args.rawdata_root:
            raise ValueError("--action init_config requires --fq_dir or --rawdata_root")
        if args.rawdata_root and not args.sample_ids:
            raise ValueError("--sample_ids is required when using --rawdata_root")
        if args.build_reference_mode == "excel":
            for field_name in ("sample_info_xlsx", "batch2_fq_dir", "batch3_fq_dir"):
                if not getattr(args, field_name):
                    raise ValueError(f"--{field_name} is required when --build_reference_mode=excel")
        if args.config_kind == "predict_samples" and args.build_reference_mode != "none":
            raise ValueError("predict_samples config must not be generated with --build_reference_mode=excel")
        init_config(
            project=args.project,
            fq_dir=args.fq_dir,
            output_config=args.output_config,
            template_snakefile=args.template_snakefile,
            config_kind=args.config_kind,
            build_reference_mode=args.build_reference_mode,
            sample_info_xlsx=args.sample_info_xlsx,
            batch2_fq_dir=args.batch2_fq_dir,
            batch3_fq_dir=args.batch3_fq_dir,
            rawdata_root=args.rawdata_root,
            sample_ids=args.sample_ids,
        )
    else:
        if not args.project:
            raise ValueError("--project is required for --action build_reference_groups")
        for field_name in ("sample_info_xlsx", "batch2_fq_dir", "batch3_fq_dir"):
            if not getattr(args, field_name):
                raise ValueError(f"--{field_name} is required for --action build_reference_groups")
        build_reference_groups_into_config(
            project=args.project,
            config_path=Path(args.output_config).resolve(),
            sample_info_xlsx=args.sample_info_xlsx,
            batch2_fq_dir=args.batch2_fq_dir,
            batch3_fq_dir=args.batch3_fq_dir,
        )


if __name__ == "__main__":
    main()
